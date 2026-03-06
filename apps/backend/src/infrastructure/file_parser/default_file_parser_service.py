import csv
import io
import json
from html.parser import HTMLParser
from typing import Callable
from xml.etree import ElementTree

from src.domain.knowledge.services import FileParserService
from src.domain.shared.exceptions import UnsupportedFileTypeError


class _HTMLTextExtractor(HTMLParser):
    _SKIP_TAGS = frozenset({"nav", "footer", "aside", "header", "script", "style"})

    def __init__(self) -> None:
        super().__init__()
        self._texts: list[str] = []
        self._skip_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in self._SKIP_TAGS:
            self._skip_depth += 1

    def handle_endtag(self, tag: str) -> None:
        if tag in self._SKIP_TAGS and self._skip_depth > 0:
            self._skip_depth -= 1

    def handle_data(self, data: str) -> None:
        if self._skip_depth == 0:
            self._texts.append(data)

    def get_text(self) -> str:
        return " ".join(self._texts).strip()


class DefaultFileParserService(FileParserService):
    def __init__(self) -> None:
        self._parsers: dict[str, Callable[[bytes], str]] = {
            "text/plain": self._parse_txt,
            "text/markdown": self._parse_txt,
            "text/csv": self._parse_csv,
            "application/json": self._parse_json,
            "text/xml": self._parse_xml,
            "application/xml": self._parse_xml,
            "text/html": self._parse_html,
            "application/pdf": self._parse_pdf,
            "application/vnd.openxmlformats-officedocument"
            ".wordprocessingml.document": self._parse_docx,
            "application/rtf": self._parse_rtf,
            "text/rtf": self._parse_rtf,
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet": self._parse_xlsx,
            "application/vnd.ms-excel": self._parse_xlsx,
        }

    def supported_types(self) -> set[str]:
        return set(self._parsers.keys())

    def parse(self, raw_bytes: bytes, content_type: str) -> str:
        parser = self._parsers.get(content_type)
        if parser is None:
            raise UnsupportedFileTypeError(content_type)
        return parser(raw_bytes)

    def _parse_txt(self, raw_bytes: bytes) -> str:
        return raw_bytes.decode("utf-8")

    def _parse_csv(self, raw_bytes: bytes) -> str:
        text = raw_bytes.decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        rows = [", ".join(row) for row in reader]
        return "\n".join(rows)

    def _parse_json(self, raw_bytes: bytes) -> str:
        data = json.loads(raw_bytes.decode("utf-8"))
        return self._extract_json_strings(data)

    def _extract_json_strings(self, data: object) -> str:
        if isinstance(data, str):
            return data
        if isinstance(data, dict):
            parts = []
            for key, value in data.items():
                parts.append(f"{key}: {self._extract_json_strings(value)}")
            return "\n".join(parts)
        if isinstance(data, list):
            parts = [self._extract_json_strings(item) for item in data]
            return "\n".join(parts)
        return str(data)

    def _parse_xml(self, raw_bytes: bytes) -> str:
        root = ElementTree.fromstring(raw_bytes.decode("utf-8"))
        texts = []
        for elem in root.iter():
            if elem.text and elem.text.strip():
                texts.append(elem.text.strip())
            if elem.tail and elem.tail.strip():
                texts.append(elem.tail.strip())
        return "\n".join(texts)

    def _parse_html(self, raw_bytes: bytes) -> str:
        extractor = _HTMLTextExtractor()
        extractor.feed(raw_bytes.decode("utf-8"))
        return extractor.get_text()

    def _parse_pdf(self, raw_bytes: bytes) -> str:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(raw_bytes))
        texts = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                texts.append(text)
        return "\n".join(texts)

    def _parse_docx(self, raw_bytes: bytes) -> str:
        from docx import Document

        doc = Document(io.BytesIO(raw_bytes))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs)

    def _parse_xlsx(self, raw_bytes: bytes) -> str:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        sheets: list[str] = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(", ".join(cells))
            if rows:
                sheets.append(f"[Sheet: {name}]\n" + "\n".join(rows))
        wb.close()
        return "\n\n".join(sheets)

    def _parse_rtf(self, raw_bytes: bytes) -> str:
        from striprtf.striprtf import rtf_to_text

        return rtf_to_text(raw_bytes.decode("utf-8"))
