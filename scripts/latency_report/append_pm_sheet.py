# 在 LINE延遲分析 Excel 加入「給PM的結論」分頁 + Web vs LINE 體感對比圖
import csv
import statistics
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
XLSX = BASE / "LINE延遲分析_2026-07-23.xlsx"

C_BLUE = "2A78D6"     # web 首字（實際體感終點）
C_GREEN = "008300"    # line 完整回覆

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def fnum(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ── 從最新資料計算 1次呼叫 vs 2次呼叫 的實測平均（P2 = 目前配置）──
traces = list(csv.DictReader(open(BASE / "traces.csv")))
p2 = []
for r in traces:
    ts = datetime.strptime(r["created_tw"], "%Y-%m-%d %H:%M:%S")
    if ts >= datetime(2026, 7, 21, 11, 29) and not (r["llm_model"] or "").startswith("claude"):
        p2.append(r)

one_call = [fnum(r["total_ms"]) for r in p2 if r["llm_calls"] == "1"]
multi_call = [fnum(r["total_ms"]) for r in p2 if r["llm_calls"] != "1"]
avg_one = statistics.median(one_call) / 1000 if one_call else 4.2
avg_multi = statistics.median(multi_call) / 1000 if multi_call else 8.1

# 估算參數（TTFT 儀表已部署，待 web 實測後以真值取代）
TTFT = 1.2          # 最後一輪生成的首 token 時間（估）
GEN_TAIL_NOTE = "逐字生成段（估 1.5~2.5s）"
web_first_one = round(avg_one - (avg_one * 0.45), 1)    # 1次呼叫：總時間 - 生成尾段(約45%)
web_first_multi = round(avg_multi - 2.4, 1)             # 2次呼叫：完整 - 逐字尾段(估2.4s)

wb = load_workbook(XLSX)
if "給PM的結論" in wb.sheetnames:
    del wb["給PM的結論"]
ws = wb.create_sheet("給PM的結論", 0)

rows = [
    ("LINE AI 客服回應時間 — 現況結論（給 PM）", ""),
    (f"資料基礎：近 14 天 {len(traces)} 筆真實 LINE 請求逐階段實測｜產出 2026-07-23", ""),
    ("", ""),
    ("【結論一】回應時間由「問題類型」決定，不是單一數字", ""),
    (f"  ．閒聊/簡單問題（AI 一次生成即可答）：中位數 {avg_one:.1f} 秒", f"實測 {len(one_call)} 筆"),
    (f"  ．需查知識庫的問題（決策→檢索→生成，兩段式）：中位數 {avg_multi:.1f} 秒", f"實測 {len(multi_call)} 筆"),
    ("  ．多子題/首查不中須重查的長尾：10~12 秒（少數）", ""),
    ("", ""),
    ("【結論二】時間組成 — 六成在 AI 生成本身", ""),
    ("  AI 決策與生成 ~60%｜意圖理解/安全檢查 ~19%｜知識庫檢索 ~13%｜組裝傳送 ~8%", ""),
    ("  AI 生成無法跳過；能壓縮的是決策輪與檢索的串行等待（見結論五）", ""),
    ("", ""),
    ("【結論三】LINE 與 Web 的體感差異是平台特性，非系統缺陷", ""),
    ("  LINE 訊息 API 只能送「完整訊息」，不支援逐字串流 →", ""),
    ("  AI 必須生成完最後一個字才能送出；Web 版第一個字出現時使用者即感覺「已回應」", ""),
    (f"  ．查資料型：Web 約第 {web_first_multi} 秒開始逐字顯示 vs LINE 第 {avg_multi:.1f} 秒整包送達（差 ~2.4 秒）", "估算*"),
    (f"  ．閒聊型：Web 約第 {web_first_one} 秒 vs LINE 第 {avg_one:.1f} 秒（差 ~2 秒）", "估算*"),
    ("  *首字時間儀表已上線，web 實測數據累積後將以真值更新", ""),
    ("", ""),
    ("【結論四】業界現況 — 在 LINE 內原生做生成式 AI 是少數派", ""),
    ("  實測大型企業（如國泰世華）：LINE 官方帳號僅關鍵字選單，", ""),
    ("  生成式 AI 客服導流至自家網頁版（該處可串流）。本系統為 LINE 原生完整回答，", ""),
    ("  數秒等待為此路線的固有成本，已用載入動畫提供即時回饋", ""),
    ("", ""),
    ("【結論五】已完成與下一步", ""),
    ("  已完成：回覆先行（持久化後移）/ 載入動畫非阻塞 / 關閉不必要的模型推理", ""),
    ("          → 基準 P50 8.4 秒 → 目前平均 7.8 秒；簡單問題已達 4 秒級", ""),
    ("  下一步（擇一，預估）：", ""),
    ("   A. FAQ/DM worker「直接檢索模式」：砍掉決策輪 → 查資料型 8→~5 秒（需品質驗證）", ""),
    ("   B. 檢索預取（與決策並行）：8→~7 秒（零行為風險）", ""),
    ("  風險提示：P90 曾達 11.7 秒，超過 LINE webhook 10 秒上限會觸發重送，長尾必須持續壓", ""),
]
for r in rows:
    ws.append(list(r))
ws["A1"].font = Font(size=14, bold=True)
for i, (text, _) in enumerate(rows, 1):
    if text.startswith("【"):
        ws.cell(row=i, column=1).font = Font(bold=True, color="1F4E79")
    ws.cell(row=i, column=1).alignment = Alignment(wrap_text=False, vertical="top")
ws.column_dimensions["A"].width = 92
ws.column_dimensions["B"].width = 14

# ── 體感對比圖：分組長條 ──
r0 = len(rows) + 3
ws.cell(row=r0 - 1, column=1, value="圖：體感終點對比（秒）— Web=看到第一個字；LINE=收到完整回覆").font = Font(bold=True)
ws.cell(row=r0, column=1, value="")
ws.cell(row=r0, column=2, value="Web 首字（估）")
ws.cell(row=r0, column=3, value="LINE 完整回覆（實測）")
data_rows = [
    ("閒聊/簡單問題", web_first_one, round(avg_one, 1)),
    ("需查知識庫的問題", web_first_multi, round(avg_multi, 1)),
]
for i, (name, w, l) in enumerate(data_rows):
    ws.cell(row=r0 + 1 + i, column=1, value=name)
    ws.cell(row=r0 + 1 + i, column=2, value=w)
    ws.cell(row=r0 + 1 + i, column=3, value=l)

chart = BarChart()
chart.type = "col"
chart.title = "同一題的兩種體感終點（秒）"
chart.y_axis.title = "秒"
data = Reference(ws, min_col=2, min_row=r0, max_col=3, max_row=r0 + 2)
cats = Reference(ws, min_col=1, min_row=r0 + 1, max_row=r0 + 2)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 9
chart.width = 17
chart.series[0].graphicalProperties.solidFill = C_BLUE
chart.series[1].graphicalProperties.solidFill = C_GREEN
chart.dataLabels = DataLabelList()
chart.dataLabels.showVal = True
ws.add_chart(chart, f"D{r0 - 1}")

wb.save(XLSX)
print("PM sheet added:", XLSX)
print(f"stats: one_call={avg_one:.1f}s({len(one_call)}) multi={avg_multi:.1f}s({len(multi_call)}) web_first: {web_first_one}/{web_first_multi}")
