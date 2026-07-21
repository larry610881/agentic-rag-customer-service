# LINE webhook 延遲分析 Excel 產生器
# 資料來源：agent_execution_traces（DB, 62 筆）+ line.webhook.timing（Cloud Run log）
import csv
import statistics
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
OUT = BASE / "LINE延遲分析_2026-07-21.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PHASE_COLORS = {
    "P0 基準": "FCE4EC",
    "P1 回覆先行": "FFF3E0",
    "P2 關閉推理": "E3F2FD",
    "P3 Haiku分級": "E8F5E9",
}


def classify(row):
    """依時間（台北時區）與模型分期。"""
    ts = datetime.strptime(row["created_tw"], "%Y-%m-%d %H:%M:%S")
    model = row["llm_model"] or ""
    if model.startswith("claude"):
        return "P3 Haiku分級"
    cutoff_deploy = datetime(2026, 7, 21, 10, 28)   # revision 00289（回覆先行）
    cutoff_none = datetime(2026, 7, 21, 11, 29)     # reasoning_effort='none' 生效
    if ts < cutoff_deploy:
        return "P0 基準"
    if ts < cutoff_none:
        return "P1 回覆先行"
    return "P2 關閉推理"


def note_for(row):
    ts = row["created_tw"]
    if ts == "2026-07-21 11:22:23":
        return "換版後首則（冷啟動，非常態）"
    if row["llm_model"].startswith("claude"):
        return "Haiku 首測（含 Anthropic 連線初始化）"
    return ""


def fnum(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def pct(values, p):
    vals = sorted(v for v in values if v is not None)
    if not vals:
        return None
    k = (len(vals) - 1) * p
    lo, hi = int(k), min(int(k) + 1, len(vals) - 1)
    return vals[lo] + (vals[hi] - vals[lo]) * (k - lo)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autow(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── 讀資料 ──
traces = list(csv.DictReader(open(BASE / "traces.csv")))
for r in traces:
    r["phase"] = classify(r)
    r["note"] = note_for(r)
    total = fnum(r["total_ms"]) or 0
    parts = [fnum(r.get(k)) or 0 for k in ("pre_agent_ms", "llm_total_ms", "tool_ms", "tail_ms")]
    r["other_ms"] = max(0, round(total - sum(parts)))

webhook = list(csv.DictReader(open(BASE / "webhook_timing.csv")))

wb = Workbook()

# ── Sheet 2: 每筆請求明細 ──
ws = wb.active
ws.title = "每筆請求明細"
cols = [
    ("時間（台北）", "created_tw", 20),
    ("優化階段", "phase", 14),
    ("總耗時(ms)", "total_ms", 12),
    ("模型", "llm_model", 18),
    ("分流 Worker", "worker", 22),
    ("前置未覆蓋(ms)\n意圖分類/守門/工具載入", "tail_ms", 16),
    ("Agent啟動前(ms)", "pre_agent_ms", 14),
    ("LLM 呼叫合計(ms)", "llm_total_ms", 15),
    ("LLM 呼叫次數", "llm_calls", 12),
    ("RAG/工具(ms)", "tool_ms", 13),
    ("其他(ms)", "other_ms", 10),
    ("輸入 tokens", "input_tokens", 12),
    ("輸出 tokens", "output_tokens", 12),
    ("備註", "note", 30),
]
ws.append([c[0] for c in cols])
style_header(ws, len(cols))
for r in traces:
    ws.append([
        (int(fnum(r.get(key)) ) if key not in ("created_tw", "phase", "llm_model", "worker", "note") and fnum(r.get(key)) is not None else (r.get(key) or ""))
        for _, key, _ in cols
    ])
for row_cells in ws.iter_rows(min_row=2):
    phase = row_cells[1].value
    color = PHASE_COLORS.get(phase)
    if color:
        row_cells[1].fill = PatternFill("solid", fgColor=color)
autow(ws, [c[2] for c in cols])
ws.freeze_panes = "A2"

# ── Sheet 3: 階段統計 ──
ws2 = wb.create_sheet("階段統計")
stat_cols = ["優化階段", "筆數", "平均總耗時(ms)", "P50(ms)", "P90(ms)", "最小(ms)", "最大(ms)",
             "平均LLM(ms)", "平均RAG/工具(ms)", "平均前置(ms)", "平均輸入tokens", "平均輸出tokens"]
ws2.append(stat_cols)
style_header(ws2, len(stat_cols))
order = ["P0 基準", "P1 回覆先行", "P2 關閉推理", "P3 Haiku分級"]
for phase in order:
    rows = [r for r in traces if r["phase"] == phase]
    if not rows:
        continue
    totals = [fnum(r["total_ms"]) for r in rows]
    def avg(key):
        vals = [fnum(r.get(key)) for r in rows if fnum(r.get(key)) is not None]
        return round(statistics.mean(vals)) if vals else None
    ws2.append([
        phase, len(rows), round(statistics.mean(totals)), round(pct(totals, 0.5)),
        round(pct(totals, 0.9)), round(min(totals)), round(max(totals)),
        avg("llm_total_ms"), avg("tool_ms"), avg("tail_ms"),
        avg("input_tokens"), avg("output_tokens"),
    ])
# 熱機版本（排除冷啟動/首測備註列）
ws2.append([])
ws2.append(["（熱機口徑：排除備註標記的冷啟動/首測樣本）"])
for phase in order:
    rows = [r for r in traces if r["phase"] == phase and not r["note"]]
    if not rows:
        continue
    totals = [fnum(r["total_ms"]) for r in rows]
    ws2.append([
        phase + "（熱機）", len(rows), round(statistics.mean(totals)),
        round(pct(totals, 0.5)), round(pct(totals, 0.9)),
        round(min(totals)), round(max(totals)), None, None, None, None, None,
    ])
autow(ws2, [18, 8, 14, 10, 10, 10, 10, 12, 14, 12, 14, 14])

# ── Sheet 4: Webhook 計時（Cloud Run log）──
ws3 = wb.create_sheet("Webhook計時")
ws3.append(["時間(UTC)", "log標示模型*", "AI處理(ms)", "LINE回覆(ms)", "持久化(ms)", "總計(ms)", "回覆字數"])
style_header(ws3, 7)
for r in webhook:
    ws3.append([r.get("timestamp", ""), r.get("llm_model", ""),
                fnum(r.get("process_message_ms")), fnum(r.get("reply_ms")),
                fnum(r.get("persist_ms")), fnum(r.get("total_ms")), fnum(r.get("answer_len"))])
ws3.append([])
ws3.append(["*log 標示模型為 bot 預設值，worker 覆寫（Haiku）不反映於此欄；實際模型以「每筆請求明細」為準。持久化(ms) 僅回覆先行版本後有值。"])
autow(ws3, [26, 16, 12, 13, 12, 10, 10])
ws3.freeze_panes = "A2"

# ── Sheet 5: 優化歷程 ──
ws4 = wb.create_sheet("優化歷程")
ws4.append(["日期（台北）", "措施", "類型", "說明", "效果"])
style_header(ws4, 5)
history = [
    ["2026-07-21 之前", "基準狀態", "—", "gpt-5.4 + 預設推理；回覆前先寫 DB；載入動畫同步等待", "p50 8.4s / p90 11.7s（52 筆）"],
    ["07-21 10:28", "回覆先行 + 動畫非阻塞", "程式碼（commit 877a0bf 系列）", "LINE 回覆提前至持久化之前；載入動畫改背景觸發", "體感 -0.4~0.6s；持久化實測僅 70~120ms 且已移出體感"],
    ["07-21 11:28", "關閉推理（reasoning=none）", "設定（DB）", "gpt-5.4 關閉思考模式（實證：其延遲主因為輸入處理，非推理）", "-0.5~1s；熱機水位 ~7s"],
    ["07-21 13:15", "模型分級（R2）", "設定（DB）", "商品查詢/門市查詢/閒聊 3 個 worker 切換至 Claude Haiku 4.5；高階客服維持 gpt-5.4", "LLM 呼叫時間約砍半（單次 ~1.4s）；首測含連線初始化，穩定水位待多筆樣本"],
    ["規劃中", "RAG 預取（與決策並行）", "程式碼", "agent 思考「要不要查」的同時先把檢索做完", "預估 -0.7~1s"],
    ["規劃中", "意圖分類提速 + 輸入瘦身", "程式碼+設定", "router 換快模型/減 context；RAG chunk 5→3 條", "預估 -0.5~1s"],
]
for h in history:
    ws4.append(h)
autow(ws4, [16, 26, 24, 52, 40])
for row_cells in ws4.iter_rows(min_row=2):
    for c in row_cells:
        c.alignment = Alignment(vertical="top", wrap_text=True)

# ── Sheet 1: 總覽（放最前面）──
ws0 = wb.create_sheet("總覽", 0)
warm = lambda ph: [fnum(r["total_ms"]) for r in traces if r["phase"] == ph and not r["note"]]
p0 = warm("P0 基準"); p2 = warm("P2 關閉推理"); p3 = warm("P3 Haiku分級")
lines = [
    ["LINE AI 客服回應時間分析報告"],
    [f"產出：2026-07-21｜資料範圍：近 14 天全部 LINE 請求（{len(traces)} 筆，逐請求逐階段實測）"],
    [],
    ["一、回應時間組成（每一則 LINE 回覆必經的階段）"],
    ["  1. 意圖理解與安全檢查（判斷問題類型、分流至對應知識庫）", "約 1~1.5 秒"],
    ["  2. AI 決策（判斷需要查詢哪些資料）", "gpt-5.4 約 1.7~3s／Haiku 約 1~1.5s"],
    ["  3. 知識庫檢索（向量搜尋 + 資料組裝）", "約 0.7~2.4 秒"],
    ["  4. AI 生成完整回答", "gpt-5.4 約 3s／Haiku 約 1.4s"],
    ["  5. 傳送回 LINE", "約 0.2 秒"],
    [],
    ["二、關鍵事實：LINE 平台限制"],
    ["  ．LINE 訊息 API 僅支援傳送「完整訊息」，不支援網頁版的逐字串流顯示"],
    ["  ．因此 AI 必須生成完整答案才能送出 — 數秒等待是所有 LINE 生成式 AI 的共同特性"],
    ["  ．實測業界：大型企業（如國泰世華）LINE 官方帳號僅提供關鍵字選單，"],
    ["    生成式 AI 客服導流至自家網頁版 — 在 LINE 內原生提供 AI 完整回答者屬少數"],
    [],
    ["三、優化成果（熱機口徑）"],
    [f"  基準（優化前）：平均 {round(statistics.mean(p0)/1000,1) if p0 else '—'} 秒｜P50 {round((pct(p0,0.5) or 0)/1000,1)} 秒｜P90 {round((pct(p0,0.9) or 0)/1000,1)} 秒（{len(p0)} 筆）"],
    [f"  目前（gpt-5.4 高階客服）：平均 {round(statistics.mean(p2)/1000,1) if p2 else '—'} 秒（{len(p2)} 筆）"],
    [f"  目前（Haiku FAQ 分級）：{('平均 ' + str(round(statistics.mean(p3)/1000,1)) + ' 秒（' + str(len(p3)) + ' 筆）') if p3 else '樣本不足，首測 7.1 秒（含一次性連線初始化）'}"],
    ["  下一輪（檢索預取 + 意圖提速）目標：4~5 秒"],
    [],
    ["四、資料說明"],
    ["  ．「每筆請求明細」= 每一則真實 LINE 訊息的逐階段耗時（毫秒）"],
    ["  ．「階段統計」= 各優化階段的平均/中位數/P90"],
    ["  ．「Webhook計時」= 伺服器端另一套獨立計時（AI 處理/回覆/寫入拆分）"],
    ["  ．「優化歷程」= 每項優化措施與其效果"],
]
for line in lines:
    ws0.append(line)
ws0["A1"].font = Font(size=14, bold=True)
for r in (4, 11, 17, 22):
    ws0.cell(row=r, column=1).font = Font(bold=True, color="1F4E79")
autow(ws0, [88, 36])

wb.save(OUT)
print("saved:", OUT)
