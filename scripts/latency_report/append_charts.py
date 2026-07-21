# 在 LINE延遲分析 Excel 加入「圖表」分頁：
# 1) 圓環圖 — 單筆請求各階段時間佔比
# 2) 長條圖 — 優化歷程平均耗時
# 3) 堆疊長條 — gpt-5.4 vs Haiku 階段組成對比
# 配色：dataviz 已驗證 categorical palette（light mode, 固定順序）
import csv
import statistics
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import PatternFillProperties  # noqa: F401 (availability check)
from openpyxl.styles import Font

BASE = Path(__file__).parent
XLSX = BASE / "LINE延遲分析_2026-07-21.xlsx"

# dataviz 驗證過的 categorical 順序（light surface）
C_BLUE = "2A78D6"    # slot1 — 前置
C_GREEN = "008300"   # slot2 — AI(LLM)
C_MAGENTA = "E87BA4" # slot3 — 檢索
C_YELLOW = "EDA100"  # slot4 — 其他
C_BLUE_LIGHT = "86B6EF"  # 藍 step250 — 規劃目標（同 hue 淺階）
STAGE_COLORS = [C_BLUE, C_GREEN, C_MAGENTA, C_YELLOW]


def fnum(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


traces = list(csv.DictReader(open(BASE / "traces.csv")))
for r in traces:
    ts = datetime.strptime(r["created_tw"], "%Y-%m-%d %H:%M:%S")
    model = r["llm_model"] or ""
    if model.startswith("claude"):
        r["phase"] = "P3"
    elif ts < datetime(2026, 7, 21, 10, 28):
        r["phase"] = "P0"
    elif ts < datetime(2026, 7, 21, 11, 29):
        r["phase"] = "P1"
    else:
        r["phase"] = "P2"


def stage_avgs(phase):
    rows = [r for r in traces if r["phase"] == phase]
    if not rows:
        return None
    def avg(key):
        vals = [fnum(x.get(key)) or 0 for x in rows]
        return statistics.mean(vals)
    total = avg("total_ms")
    front = avg("tail_ms")          # 前置未覆蓋（意圖分類/守門/載入）
    llm = avg("llm_total_ms")
    tool = avg("tool_ms")
    other = max(0.0, total - front - llm - tool)
    return {"front": front, "llm": llm, "tool": tool, "other": other, "total": total}


p0 = stage_avgs("P0")
p2 = stage_avgs("P2")
p3 = stage_avgs("P3")

wb = load_workbook(XLSX)
if "圖表" in wb.sheetnames:
    del wb["圖表"]
ws = wb.create_sheet("圖表", 1)  # 放在總覽後面

ws["A1"] = "關鍵數據圖表（資料來源：每筆請求明細，62 筆實測）"
ws["A1"].font = Font(size=13, bold=True)

# ── 資料表 1：單筆請求階段佔比（P2 關閉推理階段平均，目前 gpt-5.4 水位）──
ws["A3"] = "圖1 資料：一則回覆的時間都花在哪（P2 階段平均，毫秒）"
ws["A3"].font = Font(bold=True)
stage_names = ["意圖理解/安全檢查(前置)", "AI 決策與生成(LLM)", "知識庫檢索(RAG)", "回覆組裝/其他"]
ws.append([])  # row4 placeholder — 用直接定位寫
start = 4
for i, (name, key) in enumerate(zip(stage_names, ["front", "llm", "tool", "other"])):
    ws.cell(row=start + i, column=1, value=name)
    ws.cell(row=start + i, column=2, value=round(p2[key]))
    pct_v = p2[key] / p2["total"] * 100
    ws.cell(row=start + i, column=3, value=f"{pct_v:.0f}%")
ws.cell(row=start + 4, column=1, value="合計")
ws.cell(row=start + 4, column=2, value=round(p2["total"]))

donut = DoughnutChart()
donut.title = "一則回覆的時間組成（平均，% 佔比）"
labels = Reference(ws, min_col=1, min_row=start, max_row=start + 3)
data = Reference(ws, min_col=2, min_row=start, max_row=start + 3)
donut.add_data(data, titles_from_data=False)
donut.set_categories(labels)
donut.height = 9
donut.width = 16
donut.holeSize = 55
s = donut.series[0]
s.data_points = [
    DataPoint(idx=i, spPr=GraphicalProperties(solidFill=STAGE_COLORS[i]))
    for i in range(4)
]
donut.dataLabels = DataLabelList()
donut.dataLabels.showPercent = True
donut.dataLabels.showVal = False
ws.add_chart(donut, "E3")

# ── 資料表 2：優化歷程平均耗時（秒）──
r2 = 22
ws.cell(row=r2 - 1, column=1, value="圖2 資料：優化歷程（平均總耗時，秒）").font = Font(bold=True)
hist = [
    ("優化前基準", round(p0["total"] / 1000, 1)),
    ("回覆先行+關閉推理", round(p2["total"] / 1000, 1)),
    ("Haiku 分級(首測)", round(p3["total"] / 1000, 1) if p3 else None),
    ("下一輪目標", 4.5),
]
for i, (name, val) in enumerate(hist):
    ws.cell(row=r2 + i, column=1, value=name)
    ws.cell(row=r2 + i, column=2, value=val)

bar = BarChart()
bar.type = "col"
bar.title = "優化歷程：平均回應時間（秒）"
bar.y_axis.title = "秒"
data = Reference(ws, min_col=2, min_row=r2, max_row=r2 + 3)
cats = Reference(ws, min_col=1, min_row=r2, max_row=r2 + 3)
bar.add_data(data, titles_from_data=False)
bar.set_categories(cats)
bar.height = 9
bar.width = 16
bar.gapWidth = 60
bs = bar.series[0]
# 實測=藍、目標=淺藍（同色系淺階代表「規劃中」）
bs.data_points = [
    DataPoint(idx=0, spPr=GraphicalProperties(solidFill=C_BLUE)),
    DataPoint(idx=1, spPr=GraphicalProperties(solidFill=C_BLUE)),
    DataPoint(idx=2, spPr=GraphicalProperties(solidFill=C_BLUE)),
    DataPoint(idx=3, spPr=GraphicalProperties(solidFill=C_BLUE_LIGHT)),
]
bar.dataLabels = DataLabelList()
bar.dataLabels.showVal = True
bar.legend = None
ws.add_chart(bar, "E21")

# ── 資料表 3：gpt-5.4 vs Haiku 階段組成（堆疊，毫秒）──
r3 = 40
ws.cell(row=r3 - 1, column=1, value="圖3 資料：模型分級前後的階段組成（毫秒）").font = Font(bold=True)
ws.cell(row=r3, column=1, value="")
for j, name in enumerate(stage_names):
    ws.cell(row=r3, column=2 + j, value=name)
rows_cmp = [("gpt-5.4（高階客服）", p2), ("Claude Haiku 4.5（FAQ，首測）", p3)]
for i, (label, st) in enumerate(rows_cmp):
    ws.cell(row=r3 + 1 + i, column=1, value=label)
    if st:
        for j, key in enumerate(["front", "llm", "tool", "other"]):
            ws.cell(row=r3 + 1 + i, column=2 + j, value=round(st[key]))

stacked = BarChart()
stacked.type = "bar"       # 水平堆疊
stacked.grouping = "stacked"
stacked.overlap = 100
stacked.title = "模型分級前後：各階段耗時組成（毫秒）"
data = Reference(ws, min_col=2, min_row=r3, max_col=5, max_row=r3 + 2)
cats = Reference(ws, min_col=1, min_row=r3 + 1, max_row=r3 + 2)
stacked.add_data(data, titles_from_data=True)
stacked.set_categories(cats)
stacked.height = 8
stacked.width = 20
for i, ser in enumerate(stacked.series):
    ser.graphicalProperties.solidFill = STAGE_COLORS[i]
stacked.dataLabels = DataLabelList()
stacked.dataLabels.showVal = True
ws.add_chart(stacked, "E39")

# 註解
r4 = 58
notes = [
    "說明：",
    "．圖1 佔比為 P2 階段（目前 gpt-5.4）平均；「前置」= 意圖理解與安全檢查，為固定開銷",
    "．圖2 淺藍色為規劃目標（RAG 預取 + 意圖提速後的預估），其餘為實測平均",
    "．圖3 顯示 Haiku 將「AI 決策與生成」約砍半；「前置」與「檢索」為下一輪優化目標",
    "．Haiku 僅 1 筆樣本（含首次連線初始化），穩定水位待補測更新",
]
for i, t in enumerate(notes):
    ws.cell(row=r4 + i, column=1, value=t)

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 8
for col in "DEFG":
    ws.column_dimensions[col].width = 14

wb.save(XLSX)
print("charts added:", XLSX)
